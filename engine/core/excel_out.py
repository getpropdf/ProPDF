"""Formatted Excel output (openpyxl) with real numeric/date cells,
bank-statement layout, and a Validation sheet with continuity results."""
from __future__ import annotations

from decimal import Decimal
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .numparse import classify_cell
from .statements import StatementResult

HDR_FILL = PatternFill("solid", fgColor="1F6FEB")
HDR_FONT = Font(bold=True, color="FFFFFF")
BAD_FILL = PatternFill("solid", fgColor="FFE0E0")
THIN = Side(style="thin", color="D0D7DE")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
NUM_FMT = "#,##0.00"
DATE_FMT = "DD/MM/YYYY"


def _autowidth(ws):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[letter].width = min(60, max(9, width + 2))


def write_statement(res: StatementResult, out_path: str, source_name: str = ""):
    wb = Workbook()
    ws = wb.active
    ws.title = "Statement"
    # account-info preamble, kept like the original PDF (max 30 lines)
    for line in (res.preamble or [])[:30]:
        ws.append([line])
        ws[ws.max_row][0].font = Font(italic=True, color="555555", size=9)
    if res.preamble:
        ws.append([])
    extras = list(res.extra_headers or [])
    headers = list(res.headers) + extras + ["REVIEW"]
    ws.append(headers)
    hdr_row = ws.max_row
    for c in ws[hdr_row]:
        c.fill = HDR_FILL
        c.font = HDR_FONT
        c.border = BORDER
    bad = set(res.bad_rows)
    n_extra = len(extras)
    first_data = hdr_row + 1
    for i, r in enumerate(res.rows, start=1):
        vals = [
            r["date"], r["value_date"], r["particulars"], r["ref"],
            float(r["debit"]) if r["debit"] is not None else None,
            float(r["credit"]) if r["credit"] is not None else None,
            float(r["balance"]) if r["balance"] is not None else None,
        ]
        ex = list(r.get("extras") or [])
        ex += [""] * (n_extra - len(ex))
        vals += ex[:n_extra]
        vals.append("CHECK" if i in bad else "")
        ws.append(vals)
        row = ws[ws.max_row]
        for j, c in enumerate(row):
            c.border = BORDER
            if j in (4, 5, 6):
                c.number_format = NUM_FMT
            if j in (0, 1) and c.value is not None and not isinstance(c.value, str):
                c.number_format = DATE_FMT
            if i in bad:
                c.fill = BAD_FILL
    ws.freeze_panes = f"A{first_data}"
    n = ws.max_row
    total_row = ["", "", "TOTAL", "",
                 f"=SUM(E{first_data}:E{n})", f"=SUM(F{first_data}:F{n})", ""]
    total_row += [""] * n_extra + [""]
    ws.append(total_row)
    for c in ws[ws.max_row]:
        c.font = Font(bold=True)
        c.border = BORDER
        c.number_format = NUM_FMT
    _autowidth(ws)

    v = wb.create_sheet("Validation")
    v.append(["ProPDF Conversion Validation", source_name])
    v["A1"].font = Font(bold=True, size=13)
    rows = [
        ("Transactions extracted", len(res.rows)),
        ("Opening balance", float(res.opening_balance) if res.opening_balance is not None else "not found"),
        ("Closing balance", float(res.closing_balance) if res.closing_balance is not None else "not found"),
        ("Total withdrawals (Dr)", float(res.total_debits)),
        ("Total deposits (Cr)", float(res.total_credits)),
        ("Balance-continuity checks passed", res.continuity_ok),
        ("Balance-continuity checks FAILED", res.continuity_bad),
        ("Continuity confidence", f"{res.confidence}%"),
        ("Rows needing manual review", ", ".join(map(str, res.bad_rows)) or "none"),
    ]
    v.append([])
    for k, val in rows:
        v.append([k, val])
    if res.warnings:
        v.append([])
        v.append(["Warnings"])
        v[v.max_row][0].font = Font(bold=True)
        for w in res.warnings:
            v.append([w])
    _autowidth(v)
    wb.save(out_path)


def write_generic(grids: List[tuple], out_path: str, one_sheet: bool,
                  meta_lines: Optional[List[str]] = None):
    """grids: list of (sheet_name, grid). Cells typed via classify_cell."""
    wb = Workbook()
    first = True

    def put(ws, grid):
        # header = richest all-text row before the first numeric row
        first_num = len(grid)
        for k, row in enumerate(grid):
            if any(classify_cell(c)[0] == "number" for c in row):
                first_num = k
                break
        header_idx, header_filled = -1, 2
        for k in range(first_num):
            fc = sum(1 for c in grid[k] if str(c).strip())
            if fc > header_filled and not any(classify_cell(c)[0] == "number" for c in grid[k]):
                header_filled, header_idx = fc, k
        for ri, row in enumerate(grid):
            vals, kinds = [], []
            for cell in row:
                kind, val = classify_cell(cell)
                kinds.append(kind)
                if kind == "number":
                    vals.append(float(val))
                elif kind == "date":
                    vals.append(val)
                else:
                    vals.append(val if val != "" else None)
            ws.append(vals)
            is_hdr = ri == header_idx
            for c, kind in zip(ws[ws.max_row], kinds):
                c.border = BORDER
                if kind == "number":
                    c.number_format = NUM_FMT
                elif kind == "date":
                    c.number_format = DATE_FMT
                if is_hdr:
                    c.fill = HDR_FILL
                    c.font = HDR_FONT
        if header_idx >= 0:
            ws.freeze_panes = f"A{header_idx + 2}"
        _autowidth(ws)

    if one_sheet:
        ws = wb.active
        ws.title = "Data"
        combined = []
        for _, g in grids:
            combined.extend(g)
        put(ws, combined)
    else:
        for name, g in grids:
            ws = wb.active if first else wb.create_sheet()
            ws.title = str(name)[:31]
            first = False
            put(ws, g)

    if meta_lines:
        v = wb.create_sheet("Conversion Info")
        for line in meta_lines:
            v.append([line])
        _autowidth(v)
    wb.save(out_path)
